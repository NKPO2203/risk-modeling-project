"""
Export du tableau de l'univers retenu au format Excel.

Rassemble en un seul fichier ce qui est aujourd'hui dispersé dans quatre CSV :
la classification manuelle, la corroboration par les comptes, les informations
d'indice et l'anciennete de chaque entreprise.

Trois dates sont fournies, parce qu'elles ne disent pas la meme chose :
  Debut de cotation  premiere seance boursiere disponible
  Fondee             annee de fondation de l'entreprise
  Entree indice      date d'entree dans le S&P 500

Le debut de cotation vient de Yahoo Finance, source non officielle mais la
seule accessible gratuitement pour cette information. Il ne faut pas le lire
comme une date d'introduction en bourse dans tous les cas : quand une
entreprise nait d'une fusion ou d'un changement de nom, l'historique de prix
est herite du predecesseur. Evergy, constituee en 2018, affiche ainsi une
cotation depuis 1973, qui est celle de Westar Energy. La colonne
"Historique herite" signale ces cas, reperes par une cotation anterieure a
la fondation de l'entreprise.

Sortie : univers_82.xlsx
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RACINE = Path(__file__).resolve().parents[1]
SORTIE = RACINE / "univers_82.xlsx"

ORDRE_DEGRE = {"quasi total": 0, "fort": 1, "partiel": 2, "faible": 3}
ORDRE_CANAL = {"depense": 0, "depense et vend": 1, "vend": 2, "fournit": 3}


def construire():
    univers = pd.read_csv(RACINE / "data/processed/univers_retenu.csv")
    corro = pd.read_csv(RACINE / "data/processed/corroboration.csv")
    const = pd.read_csv(RACINE / "data/raw/sp500_constituents.csv", dtype={"CIK": str})
    faits = pd.read_csv(RACINE / "data/raw/sec_facts_raw.csv", dtype={"cik": str})
    cotations = pd.read_csv(RACINE / "data/raw/premieres_cotations.csv")

    const["CIK"] = const["CIK"].str.zfill(10)
    infos = (
        const.groupby("CIK")
        .agg(fondee=("Founded", "first"), entree_indice=("Date added", "min"))
        .reset_index()
        .rename(columns={"CIK": "cik"})
    )
    cik_par_nom = faits[["cik", "nom"]].drop_duplicates()
    t = univers.merge(cik_par_nom, on="nom", how="left")
    t = t.merge(infos, on="cik", how="left")
    t = t.merge(corro[["nom", "corroboration"]], on="nom", how="left")

    t["ticker"] = t["symboles"].str.split("|").str[0]
    t = t.merge(cotations[["ticker", "premiere_cotation"]], on="ticker", how="left")
    t["tickers"] = t["symboles"].str.replace("|", " / ", regex=False)
    t["fondee"] = t["fondee"].astype(str).str.extract(r"(\d{4})")
    an_cot = pd.to_numeric(t["premiere_cotation"].str[:4], errors="coerce")
    an_fond = pd.to_numeric(t["fondee"], errors="coerce")
    t["historique_herite"] = (an_cot < an_fond).map({True: "oui", False: ""})

    colonnes = {
        "ticker": "Ticker", "tickers": "Tickers", "nom": "Entreprise",
        "secteur": "Secteur", "sous_secteur": "Sous-secteur",
        "canal": "Canal", "degre": "Degre",
        "corroboration": "Corroboration",
        "premiere_cotation": "Debut de cotation",
        "historique_herite": "Historique herite",
        "fondee": "Fondee", "entree_indice": "Entree indice",
        "motif": "Motif retenu",
    }
    t = t[list(colonnes)].rename(columns=colonnes)
    t["rang_degre"] = t["Degre"].map(ORDRE_DEGRE)
    t["rang_canal"] = t["Canal"].map(ORDRE_CANAL)
    return t


def mettre_en_forme(ws, n_lignes):
    entete = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.fill = entete
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    largeurs = {"Ticker": 9, "Tickers": 13, "Entreprise": 30, "Secteur": 24,
                "Sous-secteur": 34, "Canal": 17, "Degre": 12, "Corroboration": 17,
                "Debut de cotation": 15, "Historique herite": 13,
                "Entree indice": 13, "Fondee": 9, "Motif retenu": 95}
    for i, c in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeurs.get(c.value, 14)
    for ligne in ws.iter_rows(min_row=2, max_row=n_lignes + 1):
        for c in ligne:
            c.alignment = Alignment(vertical="top", wrap_text=(c.column_letter == "O"))


def main():
    t = construire()
    cols = [c for c in t.columns if not c.startswith("rang_")]

    par_date = t.sort_values("Debut de cotation", ascending=False)[cols]
    par_secteur = t.sort_values(
        ["Secteur", "rang_degre", "Debut de cotation"], ascending=[True, True, False])[cols]
    par_canal = t.sort_values(
        ["rang_canal", "rang_degre", "Entreprise"])[cols]

    recap = (
        t.groupby(["Secteur", "Degre"]).size().reset_index(name="Nombre")
        .sort_values(["Secteur", "Degre"])
    )

    with pd.ExcelWriter(SORTIE, engine="openpyxl") as w:
        par_date.to_excel(w, sheet_name="Par anciennete", index=False)
        par_secteur.to_excel(w, sheet_name="Par secteur", index=False)
        par_canal.to_excel(w, sheet_name="Par canal", index=False)
        recap.to_excel(w, sheet_name="Recapitulatif", index=False)
        for nom, df in [("Par anciennete", par_date), ("Par secteur", par_secteur),
                        ("Par canal", par_canal), ("Recapitulatif", recap)]:
            mettre_en_forme(w.sheets[nom], len(df))

    print(f"{len(t)} entreprises ecrites dans {SORTIE.name}")
    print("\nles 14 cotations les plus recentes :")
    for _, r in par_date.head(14).iterrows():
        print(f"  {r['Debut de cotation']}  {r['Ticker']:6s} {r['Entreprise'][:28]:28s} {r['Secteur'][:20]}")
    herites = t[t["Historique herite"] == "oui"]
    print(f"\n{len(herites)} entreprises a historique herite : {', '.join(herites['Ticker'])}")


if __name__ == "__main__":
    main()
